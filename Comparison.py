import time
import pandas as pd
import numpy as np
import matplotlib.pylab as plt
import seaborn as sns
import pulp as p
import random
from itertools import combinations
import winsound
import os
import openpyxl

main_dir = "Images"
os.mkdir(main_dir)

# Call a Workbook() function of openpyxl to create a new blank Workbook object
wb = openpyxl.Workbook()
# Get workbook active sheet from the active attribute
sheet = wb.active
row_number_on_Excel_Sheet=1
cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 1)
cell.value = "Number of Nodes used (All Vehicles have been considered)"
cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 2)
cell.value = "Optimal Objective Value WITH Constraints 9 and 10"
cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 4)
cell.value = "Optimal Objective Value WITHOUT Constraints 9 and 10"
cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 3)
cell.value = "Time Taken for Solver to compute WITH Constraints 9 and 10"
cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 5)
cell.value = "Time Taken for Solver to compute WITHOUT Constraints 9 and 10"

# Get the Input
Nodes=pd.read_excel("Input Data.xlsx","Locations & Delivery-PickUp",index_col=0)
range_of_number_of_nodes=Nodes.shape[0]

Vehicles=pd.read_excel("Input Data.xlsx","Vehicle Specifications",index_col=0)
num_of_Vehicle_types=Vehicles.shape[0]

for num_of_Nodes in range(1,range_of_number_of_nodes+1):

    row_number_on_Excel_Sheet+=1
    cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 1)
    cell.value =num_of_Nodes
    wb.save("Comparison Table.xlsx")

    main_dir = "Images/"+str(num_of_Nodes)
    os.mkdir(main_dir)

    #Sets Used
    Relief_Centres=set()
    Depot_and_Relief_Centres=set()

    counter=0
    for i, row in Nodes.iterrows():
        if i!=0: # PLEASE ENSURE 0 IS DEFINED AS THE DEPOT. ALSO ENSURE THE SOLO DEPOT HAS NODE NUMBER 0.
            Relief_Centres.add(i)
        Depot_and_Relief_Centres.add(i)
        counter+=1
        if counter>num_of_Nodes:
            break

    counter=0
    Vehicle_Types=set()
    VN={}
    VQ={}
    VS={}
    VC={}
    for i, row in Vehicles.iterrows():
        Vehicle_Types.add(i)
        VN[i]=row["VN"]
        VQ[i]=row["VQ"]
        VS[i]=row["VS"]
        VC[i]=row["VC"]
        counter+=1
        if counter>=num_of_Vehicle_types:
            break

    pp={}
    d={}
    for i, row in Nodes.iterrows():
        pp[i]=row["PickUp"]
        d[i]=row["Delivery"]
        if i==num_of_Nodes:
            break

    #Creating the Distance Matrices
    C={} #This is the COST/DISTANCE Matrix

    #For checking the exact solution of the paper here all the distances are considered Euclidean
    for k in Vehicle_Types:
        temp = pd.read_excel("Input Data.xlsx","Calculating Random Distances")
        for i,row in temp.iterrows():
            if row["Origin Node"] in Depot_and_Relief_Centres:
                if row["Destination Node"] in Depot_and_Relief_Centres:
                    key=(int(row["Origin Node"]),int(row["Destination Node"]),int(k))
                    value=float(row["Euclidean Distance"])
                    C[key]=value
    """
    for k in Vehicle_Types:
        temp = pd.read_excel("Input Data.xlsx",str(k))
        for i,row in temp.iterrows():
            if row["Origin Node"] in Depot_and_Relief_Centres:
                if row["Destination Node"] in Depot_and_Relief_Centres:
                    key=(int(row["Origin Node"]),int(row["Destination Node"]),int(k))
                    value=float(row["Distance"])
                    C[key]=value
    """

    # Set the problem
    prob=p.LpProblem("Heterogenenous_single_Depot_mVRPSDP",p.LpMinimize)

    # Decision Variables
    # Iff Arc joining i & j is included within the solution for the Layer k
    x=p.LpVariable.dicts('x',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types),cat='Binary')

    # Amount of collected load across Arc(i,j) by a Vehicle in Layer k
    y=p.LpVariable.dicts('y',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types),lowBound=0)

    # Amount of delivery load across Arc(i,j) done by a Vehicle in Layer k
    z=p.LpVariable.dicts('z',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types),lowBound=0)

    # Set Objective Function (Point 2)
    prob+=p.lpSum(x[i,j,k]*C[i,j,k]*VS[k] for k in Vehicle_Types for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres)+p.lpSum(x[0,j,k]*VC[k] for k in Vehicle_Types for j in Relief_Centres)

    #Ensuring at most a single vehicle caters to a Relief Center (Point 3 a)
    for i in Relief_Centres:
        prob+=p.lpSum(x[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types)<=1

    # Ensuring equal number of Incoming and Outgoing paths are available from all Nodes (Point 3 b)
    for i in Depot_and_Relief_Centres:
        for k in Vehicle_Types:
            prob+=p.lpSum(x[i,j,k]-x[j,i,k] for j in Depot_and_Relief_Centres)==0

    # Ensuring at most VN outgoing paths are available at the Depot since there are VN[k] vehicle for each Vehicle Type (Point 3 c)
    for k in Vehicle_Types:
        prob+=p.lpSum(x[0,j,k] for j in Relief_Centres)<=VN[k]

    """Flow Limitation Constraints"""

    for k in Vehicle_Types:
        for j in Relief_Centres:
            prob+=y[0,j,k]==0   #Ensuring initial PickUp from Nodes is 0 (Point 3 d i)

    for k in Vehicle_Types:
        for i in Relief_Centres:
            prob+=z[i,0,k]==0   #Ensuring final Delivery to Nodes is 0 (Point 3 d ii)

    #Ensuring the PickUp constraints are satisfied (Point 3 e i)
    for i in Relief_Centres:
        prob+=p.lpSum(y[i,j,k]-y[j,i,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types)==pp[i]

    #Ensuring the Delivery constraints are satisfied (Point 3 e ii)
    for i in Relief_Centres:
        prob+=p.lpSum(z[j,i,k]-z[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types)==d[i]

    """Constraining the Sum of Flows to and from the Origin/Depot/Warehouse/NDRF_BASE"""
    #(Point 3 f i)
    prob+=p.lpSum(y[i,0,k] for i in Relief_Centres for k in Vehicle_Types)==p.lpSum(pp[i] for i in Relief_Centres)   #Ensuring sum of all PickUp Flow Variables to the Origin [0th Node] is equal to the total PickUps of all Nodes
    #(Point 3 f ii)
    prob+=p.lpSum(z[0,i,k] for i in Relief_Centres for k in Vehicle_Types)==p.lpSum(d[i] for i in Relief_Centres)   #Ensuring sum of all Delivery Flow Variables from the Origin [0th Node] is equal to the total Demand of all Nodes

    # Ensuring the vehicle capacity is never exceeded (Point 3 f)
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                prob+=y[i,j,k]+z[i,j,k]<=VQ[k]*x[i,j,k]

    # Solve the Problem using default CBC
    start_time=time.time()
    #status=prob.solve(p.PULP_CBC_CMD(maxSeconds=300, msg=1, gapRel=0))
    #status=prob.solve(p.PULP_CBC_CMD(timeLimit=700))
    status=prob.solve()
    #status=prob.solve(p.PULP_CBC_CMD(timeLimit=999))
    end_time=time.time()

    winsound.Beep(555-19*num_of_Nodes, 888+19*num_of_Nodes) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
    print("This is the status:- ", p.LpStatus[prob.status])
    objec_val=p.value(prob.objective)

    main_dir=main_dir+"/{}"
    # Draw the optimal routes Layerwise
    for k in Vehicle_Types:
        plt.figure(figsize=(9,9))
        for i, row in Nodes.iterrows():
            if i>num_of_Nodes:
                break
            if i==0:
                plt.scatter(row["Longitude"],row["Latitude"], c='r',marker='s')
                plt.text( row["Longitude"] + 0.33, row["Latitude"] + 0.33, "Depot")
            else:
                plt.scatter(row["Longitude"], row["Latitude"], c='black')
                plt.text(row["Longitude"] + 0.33, row["Latitude"] + 0.33, i)
        plt.title('mVRPSDC Tours for Vehicles of Type '+str(k)+" on the corresponding layer "+str(k))
        plt.ylabel("Latitude")
        plt.xlabel("Longitude")

        max=0   # Finding the maximum utilised vehicle capacity
        routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if p.value(x[i,j,k])==1]
        arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor='blue')
        for i, j in routes:
            utilized_capacity=y[i,j,k].varValue+z[i,j,k].varValue
            if utilized_capacity>max:
                max=utilized_capacity
            plt.annotate('', xy=[Nodes.iloc[j]["Longitude"], Nodes.iloc[j]["Latitude"]], xytext=[Nodes.iloc[i]["Longitude"], Nodes.iloc[i]["Latitude"]], arrowprops=arrowprops)    
            #plt.text((Nodes.iloc[i]["Longitude"]+Nodes.iloc[j]["Longitude"])/2, (Nodes.iloc[i]["Latitude"]+Nodes.iloc[j]["Latitude"])/2, f'{utilized_capacity}',fontweight="bold")

        print("The maximum vehicle capacity utilised ever in any tour in layer ",k," is: ",max," out of the total available",VQ[k])

        used_vehicles=0 # Finding the maximum number of vehicles being used
        for j in Relief_Centres:
            used_vehicles=p.value(x[0,j,k])+used_vehicles
        print("The maximum numbers of vehicles used is: ",used_vehicles," out of total available ",VN[k])
        name="With the Original Constraints of 9 and 10, Vehicles_ "+str(used_vehicles)+"--"+str(VN[k])+" and Capacity_ "+str(max)+"--"+str(VQ[k])+" with Objective Value_ "+str(objec_val)+" & Solver Time is_ "+str(end_time-start_time)+"seconds.png"
        plt.savefig(main_dir.format(name))

    cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 2)
    cell.value = objec_val
    cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 3)
    cell.value = end_time-start_time
    wb.save("Comparison Table.xlsx")


    # Set the problem
    prob=p.LpProblem("Heterogenenous_single_Depot_mVRPSDP",p.LpMinimize)

    # Decision Variables
    # Iff Arc joining i & j is included within the solution for the Layer k
    x=p.LpVariable.dicts('x',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types),cat='Binary')

    # Amount of collected load across Arc(i,j) by a Vehicle in Layer k
    y=p.LpVariable.dicts('y',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types),lowBound=0)

    # Amount of delivery load across Arc(i,j) done by a Vehicle in Layer k
    z=p.LpVariable.dicts('z',((i,j,k) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres for k in Vehicle_Types),lowBound=0)

    # Set Objective Function (Point 2)
    prob+=p.lpSum(x[i,j,k]*C[i,j,k]*VS[k] for k in Vehicle_Types for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres)+p.lpSum(x[0,j,k]*VC[k] for k in Vehicle_Types for j in Relief_Centres)

    #Ensuring at most a single vehicle caters to a Relief Center (Point 3 a)
    for i in Relief_Centres:
        prob+=p.lpSum(x[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types)<=1

    # Ensuring equal number of Incoming and Outgoing paths are available from all Nodes (Point 3 b)
    for i in Depot_and_Relief_Centres:
        for k in Vehicle_Types:
            prob+=p.lpSum(x[i,j,k]-x[j,i,k] for j in Depot_and_Relief_Centres)==0

    # Ensuring at most VN outgoing paths are available at the Depot since there are VN[k] vehicle for each Vehicle Type (Point 3 c)
    for k in Vehicle_Types:
        prob+=p.lpSum(x[0,j,k] for j in Relief_Centres)<=VN[k]

    """Flow Limitation Constraints"""

    for k in Vehicle_Types:
        for j in Relief_Centres:
            prob+=y[0,j,k]==0   #Ensuring initial PickUp from Nodes is 0 (Point 3 d i)

    for k in Vehicle_Types:
        for i in Relief_Centres:
            prob+=z[i,0,k]==0   #Ensuring final Delivery to Nodes is 0 (Point 3 d ii)

    #Ensuring the PickUp constraints are satisfied (Point 3 e i)
    for i in Relief_Centres:
        prob+=p.lpSum(y[i,j,k]-y[j,i,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types)==pp[i]

    #Ensuring the Delivery constraints are satisfied (Point 3 e ii)
    for i in Relief_Centres:
        prob+=p.lpSum(z[j,i,k]-z[i,j,k] for j in Depot_and_Relief_Centres for k in Vehicle_Types)==d[i]

    # Ensuring the vehicle capacity is never exceeded (Point 3 f)
    for i in Depot_and_Relief_Centres:
        for j in Depot_and_Relief_Centres:
            for k in Vehicle_Types:
                prob+=y[i,j,k]+z[i,j,k]<=VQ[k]*x[i,j,k]

    # Solve the Problem using default CBC
    start_time_without_the_constraints=time.time()
    #status=prob.solve(p.PULP_CBC_CMD(maxSeconds=300, msg=1, gapRel=0))
    #status=prob.solve(p.PULP_CBC_CMD(timeLimit=700))
    status=prob.solve()
    #status=prob.solve(p.PULP_CBC_CMD(timeLimit=999))
    end_time_without_the_constraints=time.time()

    winsound.Beep(555-19*num_of_Nodes, 888+19*num_of_Nodes) # where 500 is the frequency in Hertz and 1000 is the duration in miliseconds
    print("This is the status:- ", p.LpStatus[prob.status])
    objec_val_without_constraints=p.value(prob.objective)

    # Draw the optimal routes Layerwise
    for k in Vehicle_Types:
        plt.figure(figsize=(9,9))
        for i, row in Nodes.iterrows():
            if i>num_of_Nodes:
                break
            if i==0:
                plt.scatter(row["Longitude"],row["Latitude"], c='r',marker='s')
                plt.text( row["Longitude"] + 0.33, row["Latitude"] + 0.33, "Depot")
            else:
                plt.scatter(row["Longitude"], row["Latitude"], c='black')
                plt.text(row["Longitude"] + 0.33, row["Latitude"] + 0.33, i)
        plt.title('mVRPSDC Tours for Vehicles of Type '+str(k)+" on the corresponding layer "+str(k))
        plt.ylabel("Latitude")
        plt.xlabel("Longitude")

        max=0   # Finding the maximum utilised vehicle capacity
        routes = [(i, j) for i in Depot_and_Relief_Centres for j in Depot_and_Relief_Centres if p.value(x[i,j,k])==1]
        arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor='blue')
        for i, j in routes:
            utilized_capacity=y[i,j,k].varValue+z[i,j,k].varValue
            if utilized_capacity>max:
                max=utilized_capacity
            plt.annotate('', xy=[Nodes.iloc[j]["Longitude"], Nodes.iloc[j]["Latitude"]], xytext=[Nodes.iloc[i]["Longitude"], Nodes.iloc[i]["Latitude"]], arrowprops=arrowprops)    
            #plt.text((Nodes.iloc[i]["Longitude"]+Nodes.iloc[j]["Longitude"])/2, (Nodes.iloc[i]["Latitude"]+Nodes.iloc[j]["Latitude"])/2, f'{utilized_capacity}',fontweight="bold")

        print("The maximum vehicle capacity utilised ever in any tour in layer ",k," is: ",max," out of the total available",VQ[k])

        used_vehicles=0 # Finding the maximum number of vehicles being used
        for j in Relief_Centres:
            used_vehicles=p.value(x[0,j,k])+used_vehicles
        print("The maximum numbers of vehicles used is: ",used_vehicles," out of total available ",VN[k])
        name="Vehicles_ "+str(used_vehicles)+"--"+str(VN[k])+" and Capacity_ "+str(max)+"--"+str(VQ[k])+" with Objective Value_ "+str(objec_val_without_constraints)+" & Solver Time is_ "+str(end_time_without_the_constraints-start_time_without_the_constraints)+"seconds.png"
        plt.savefig(main_dir.format(name))

    
    cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 4)
    cell.value = objec_val_without_constraints
    cell = sheet.cell(row = row_number_on_Excel_Sheet, column = 5)
    cell.value = end_time_without_the_constraints-start_time_without_the_constraints
    wb.save("Comparison Table.xlsx")